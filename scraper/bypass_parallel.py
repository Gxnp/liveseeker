# bypass_parallel.py
"""
Parallel m3u8 scanner (preserves play-click + iframe/shadow handling + refresh-for-dooball)
- Each worker opens its own Chrome (selenium-wire) instance
- Non-dooball: open/close per visit (fresh session) -> implemented as "batched rounds"
- dooball: open once (actual_visits = 1), run refresh-click loop to collect multiple m3u8
- driver.scopes set to catch only .m3u8
"""
import time
import random
import json
import re
import os
from datetime import datetime
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Set, Dict, Tuple, List

import chromedriver_autoinstaller
from seleniumwire import webdriver  # pip install selenium-wire
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import (
	ElementClickInterceptedException,
	ElementNotInteractableException,
	StaleElementReferenceException,
	JavascriptException
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------
# CONFIG
# -----------------------------
SITES = [
""
]
VISITS_PER_SITE = 8
SELECTORS_FILE = "./selectors.json"
M3U8_ONLY_SCOPES = True
RESULTS_FOLDER = "./results"  # folder สำหรับเก็บผลลัพธ์

# Parallel config
MAX_WORKERS = 5   #max browser
HEADLESS = False    # set False for debugging / visual
# play-wait tuning (smaller for speed, increase if unreliable)
PLAY_WAIT_MIN = 0.8
PLAY_WAIT_MAX = 1.2

# -----------------------------
# Excel helpers
# -----------------------------
def export_xlsx(result_m3u8_map: Dict[str, Set[str]], filename=None):
	# สร้าง folder results ถ้ายังไม่มี
	os.makedirs(RESULTS_FOLDER, exist_ok=True)
	
	# ถ้าไม่ระบุ filename ให้สร้างชื่อไฟล์ด้วย timestamp
	if filename is None:
		timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
		filename = f"{timestamp}_m3u8.xlsx"
	
	# รวม path ของ folder กับ filename
	filepath = os.path.join(RESULTS_FOLDER, filename)
	
	wb = Workbook()
	ws = wb.active
	ws.title = "m3u8_links"
	ws.append(["scanned_site", "m3u8_urls"])
	for site, links in result_m3u8_map.items():
		ws.append([site, "\n".join(sorted(links))])
	for cell in ws[1]:
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center")
	for col in range(1, ws.max_column + 1):
		col_letter = get_column_letter(col)
		max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
		ws.column_dimensions[col_letter].width = min(max_len + 2, 64)
	wb.save(filepath)
	print(f"✅ Exported result to {filepath}")

# -----------------------------
# utils: human-like motion + selectors load
# -----------------------------
def human_pause(a=0.12, b=0.42):
	time.sleep(random.uniform(a, b))

def human_pause_long(a=0.6, b=1.8):
	time.sleep(random.uniform(a, b))

def human_wiggle(driver, element, radius=6, steps=3):
	try:
		actions = ActionChains(driver)
		actions.move_to_element_with_offset(element, 1, 1).perform()
		for _ in range(steps):
			dx = random.randint(-radius, radius)
			dy = random.randint(-radius, radius)
			actions.move_by_offset(dx, dy).perform()
			time.sleep(random.uniform(0.02, 0.08))
		actions.move_to_element(element).perform()
	except Exception:
		pass

def scroll_into_view(driver, element, center=True):
	try:
		driver.execute_script("arguments[0].scrollIntoView({block: arguments[1]});", element, "center" if center else "nearest")
		human_pause(0.08, 0.18)
	except Exception:
		try:
			driver.execute_script("""
				const r = arguments[0].getBoundingClientRect();
				window.scrollBy(0, r.top - (window.innerHeight/2));
			""", element)
		except Exception:
			pass

def safe_click(driver, element, allow_js_fallback=True):
	if element is None:
		return False
	try:
		human_wiggle(driver, element)
		element.click()
		return True
	except (ElementClickInterceptedException, ElementNotInteractableException, StaleElementReferenceException):
		if allow_js_fallback:
			try:
				driver.execute_script("arguments[0].click();", element)
				return True
			except JavascriptException:
				return False
		return False

def load_selectors(path=SELECTORS_FILE):
	try:
		with open(path, "r", encoding="utf-8") as f:
			return json.load(f)
	except Exception:
		return {}

# -----------------------------
# network capture helpers
# -----------------------------
def normalize_url(url):
	try:
		parsed = urlparse(url)
		host = (parsed.hostname or "").lower()
		normalized_path = f"{parsed.scheme}://{parsed.netloc}{parsed.path}".lower()
		return normalized_path, host
	except Exception:
		return url.lower().split("?")[0], ""

def is_m3u8(url: str) -> bool:
	u, _ = normalize_url(url)
	return u.endswith(".m3u8")

def capture_network(driver):
	found = []
	for request in getattr(driver, "requests", []):
		try:
			u = request.url
			if u and is_m3u8(u):
				found.append(u)
		except Exception:
			continue
	return found

# -----------------------------
# dooball aggressive skip helper
# -----------------------------
def ensure_stream_start(driver):
	def click_video():
		vids = driver.find_elements(By.TAG_NAME, "video")
		for v in vids:
			try:
				scroll_into_view(driver, v)
				ActionChains(driver).move_to_element(v).click().perform()
				human_pause(0.6, 1.0)
				return True
			except Exception:
				continue
		return False

	# 1) main page
	try:
		body = driver.find_element(By.TAG_NAME, "body")
		ActionChains(driver).move_to_element(body).click().perform()
		human_pause(0.3, 0.6)
	except Exception:
		pass

	if click_video():
		return True

	# 2) iframe
	iframes = driver.find_elements(By.TAG_NAME, "iframe")
	for f in iframes:
		try:
			driver.switch_to.frame(f)
			if click_video():
				driver.switch_to.default_content()
				return True
			driver.switch_to.default_content()
		except Exception:
			try:
				driver.switch_to.default_content()
			except Exception:
				pass

	return False

def for_each_context(driver):
	yield None  # main page
	iframes = driver.find_elements(By.TAG_NAME, "iframe")
	for iframe in iframes:
		yield iframe

def force_skip_via_js(driver):
	try:
		driver.execute_script("""
			window.canSkipAd = true;
			window.skipAvailable = true;
			window.adFinished = true;
		""")
		return True
	except Exception:
		return False

def handle_skip_ads_dooball(driver, selectors, rounds=2):
	"""
	Aggressive skip ads
	- main page + all iframes
	- force JS + enable + click
	"""
	for _ in range(rounds):
		for iframe in for_each_context(driver):
			try:
				if iframe:
					driver.switch_to.frame(iframe)

				force_skip_via_js(driver)

				for sel in selectors.get("skip_ads_button", []):
					elements = find_elements_by_selector(driver, sel)
					for el in elements:
						try:
							enable_and_click(driver, el)
						except Exception:
							continue

				if iframe:
					driver.switch_to.default_content()
			except Exception:
				try:
					driver.switch_to.default_content()
				except Exception:
					pass

		time.sleep(0.6)

# -----------------------------
# iframe / shadow helpers
# -----------------------------
def activate_player(driver):
	iframes = []

	try:
		iframes = driver.find_elements(By.TAG_NAME, "iframe")
		for f in iframes:
			try:
				scroll_into_view(driver, f)
				driver.execute_script("arguments[0].click();", f)
				human_pause(0.3, 0.6)
			except Exception:
				pass
	except Exception:
		pass

	for f in iframes:
		try:
			driver.switch_to.frame(f)

			# click body
			try:
				body = driver.find_element(By.TAG_NAME, "body")
				ActionChains(driver).move_to_element(body).click().perform()
				human_pause(0.2, 0.4)
			except Exception:
				pass

			# click video elements
			try:
				vids = driver.find_elements(By.TAG_NAME, "video")
				for v in vids:
					safe_click(driver, v)
			except Exception:
				pass

			driver.switch_to.default_content()
		except Exception:
			try:
				driver.switch_to.default_content()
			except Exception:
				pass

def find_elements_by_selector(driver, sel):
	found = []

	sel_type = sel.get("type")
	value = sel.get("value")

	try:
		if sel_type == "css":
			found = driver.find_elements(By.CSS_SELECTOR, value)

		elif sel_type == "xpath":
			found = driver.find_elements(By.XPATH, value)

		elif sel_type == "id":
			el = driver.find_element(By.ID, value)
			found = [el] if el else []

		elif sel_type == "js":
			el = driver.execute_script(f"return {value};")
			if el:
				found = [el]

		elif sel_type == "keyword":
			keyword = value.lower()
			candidates = driver.find_elements(By.XPATH, "//*[self::button or self::div or self::span]")
			for el in candidates:
				try:
					if keyword in (el.text or "").lower():
						found.append(el)
				except Exception:
					continue
	except Exception:
		pass

	return found

def enable_and_click(driver, el):
	try:
		driver.execute_script("""
			arguments[0].disabled = false;
			arguments[0].removeAttribute('disabled');
			arguments[0].style.pointerEvents = 'auto';
			arguments[0].style.opacity = '1';
		""", el)
	except Exception:
		pass

	return safe_click(driver, el)

def try_skip_in_current_context(driver, selectors, max_wait=10):
	skip_selectors = selectors.get("skip_ads_button", [])
	clicked_any = False

	for sel in skip_selectors:
		elements = find_elements_by_selector(driver, sel)

		for el in elements:
			try:
				text = el.text or ""

				# ถ้ามี countdown ในข้อความ
				m = re.search(r"(\d+)", text)
				if m:
					wait_time = min(int(m.group(1)), max_wait)
					time.sleep(wait_time)

				# ถ้า disabled → enable
				try:
					if el.get_attribute("disabled"):
						enable_and_click(driver, el)
					else:
						safe_click(driver, el)
				except Exception:
					enable_and_click(driver, el)

				human_pause(0.4, 0.8)
				safe_click(driver, el)  # click ซ้ำ กัน delay

				clicked_any = True
			except Exception:
				continue

	return clicked_any

def handle_skip_ads(driver, selectors, iframe_depth=2):
	"""
	- ทำ skip ads ในหน้าเว็บหลัก
	- แล้ววนเข้า iframe ทุกตัว (รองรับซ้อน)
	"""
	# 1️⃣ หน้าเว็บหลัก
	try:
		try_skip_in_current_context(driver, selectors)
	except Exception:
		pass

	# 2️⃣ iframe
	def recurse_iframe(depth):
		if depth <= 0:
			return

		iframes = driver.find_elements(By.TAG_NAME, "iframe")
		for iframe in iframes:
			try:
				driver.switch_to.frame(iframe)
				try_skip_in_current_context(driver, selectors)
				recurse_iframe(depth - 1)
				driver.switch_to.parent_frame()
			except Exception:
				try:
					driver.switch_to.parent_frame()
				except Exception:
					pass

	recurse_iframe(iframe_depth)

def try_switch_to_any_iframe(driver):
	iframes = driver.find_elements(By.TAG_NAME, "iframe")
	for f in iframes:
		try:
			driver.switch_to.frame(f)
			return f
		except Exception:
			continue
	return None

def switch_back_to_default(driver):
	try:
		driver.switch_to.default_content()
	except Exception:
		pass

def attempt_click_in_shadow(driver, host_selector, inner_selector):
	"""
	Returns True if clicked inside shadow DOM or host element found & clicked.
	"""
	try:
		js = f"""
		const host = document.querySelector("{host_selector}");
		if (!host) return null;
		const root = host.shadowRoot || host;
		const el = root.querySelector("{inner_selector}");
		return el;
		"""
		el = driver.execute_script(js)
		if el:
			try:
				driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
			except Exception:
				pass
			try:
				driver.execute_script("arguments[0].click();", el)
				return True
			except Exception:
				return False
	except Exception:
		pass
	return False

# -----------------------------
# click_media_play_button (comprehensive)
# -----------------------------
def click_media_play_button(driver, selectors: dict, timeout=10) -> bool:
	"""
	Try multiple strategies to start the live player:
	1) direct visible buttons containing 'play'
	2) click <video> elements
	3) try inside iframes (switch into each iframe and repeat)
	4) shadow DOM media-player attempts
	Returns True if any click succeeded.
	"""
	# 1) direct 'play' button (text/aria)
	try:
		btns = driver.find_elements(By.XPATH,
			"//button[contains(translate(@aria-label,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'play') or "
			"contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'play')]")
		for b in btns:
			try:
				if safe_click(driver, b):
					human_pause_long(PLAY_WAIT_MIN, PLAY_WAIT_MAX)
					return True
			except Exception:
				continue
	except Exception:
		pass

	# 2) video elements on page
	try:
		vids = driver.find_elements(By.TAG_NAME, "video")
		for v in vids:
			try:
				scroll_into_view(driver, v)
				if safe_click(driver, v):
					human_pause_long(PLAY_WAIT_MIN, PLAY_WAIT_MAX)
					return True
			except Exception:
				continue
	except Exception:
		pass

	# 3) try inside each iframe (some players live in nested frames)
	iframes = driver.find_elements(By.TAG_NAME, "iframe")
	for f in iframes:
		try:
			driver.switch_to.frame(f)
			# try video elements inside iframe
			try:
				vids = driver.find_elements(By.TAG_NAME, "video")
				for v in vids:
					try:
						scroll_into_view(driver, v)
						if safe_click(driver, v):
							switch_back_to_default(driver)
							human_pause_long(PLAY_WAIT_MIN, PLAY_WAIT_MAX)
							return True
					except Exception:
						continue
			except Exception:
				pass

			# try play buttons inside iframe text
			try:
				btns = driver.find_elements(By.XPATH,
					"//button[contains(translate(@aria-label,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'play') or "
					"contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'play')]")
				for b in btns:
					try:
						if safe_click(driver, b):
							switch_back_to_default(driver)
							human_pause_long(PLAY_WAIT_MIN, PLAY_WAIT_MAX)
							return True
					except Exception:
						continue
			except Exception:
				pass

			switch_back_to_default(driver)
		except Exception:
			# couldn't switch to this frame -> continue
			try:
				switch_back_to_default(driver)
			except Exception:
				pass

	# 4) shadow DOM custom elements (media-player)
	try:
		hosts = driver.find_elements(By.CSS_SELECTOR, "media-player")
		for host in hosts:
			ok = attempt_click_in_shadow(driver, "media-player", "media-play-button[aria-label='Play']")
			if ok:
				human_pause_long(PLAY_WAIT_MIN, PLAY_WAIT_MAX)
				return True
	except Exception:
		pass

	return False

# -----------------------------
# refresh channels (Modified: Re-play + Re-skip ads after click)
# -----------------------------
def click_refresh_channels(driver, selectors: dict, already_found_links: Set[str], rounds: int = 6, delay: int = 6):
	"""
	selectors expected to contain "refresh_buttons": list of {"type":"css"/"xpath"/"id"/"js","value":...}
	Behavior:
	  - iterate rounds
	  - click refresh button -> WAIT -> CLICK PLAY -> SKIP ADS -> CAPTURE NETWORK
	"""
	refresh_buttons = selectors.get("refresh_buttons", [])
	if not refresh_buttons:
		print("[WARN] no refresh_buttons in selectors.json")
		return

	for r in range(rounds):
		print(f"[refresh] round {r+1}/{rounds}")
		for btn in refresh_buttons:
			btn_type = btn.get("type")
			btn_value = btn.get("value")
			el = None
			try:
				if btn_type == "css":
					el = driver.find_element(By.CSS_SELECTOR, btn_value)
				elif btn_type == "xpath":
					el = driver.find_element(By.XPATH, btn_value)
				elif btn_type == "id":
					el = driver.find_element(By.ID, btn_value)
				elif btn_type == "js":
					try:
						el = driver.execute_script(f"return {btn_value};")
					except Exception:
						el = None
				else:
					el = None
			except Exception:
				el = None

			if not el:
				print(f"[refresh] target not found: {btn_value}")
				human_pause(0.2, 0.5)
				continue

			# clear requests buffer then click
			try:
				# best-effort clear requests list
				try:
					del driver.requests[:]
				except Exception:
					pass
				
				driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
				
				print(f"    -> Clicking refresh/channel button...")
				safe_click(driver, el)
				

				# หลังเปลี่ยนช่อง
				activate_player(driver)
				ensure_stream_start(driver)

				# --- ส่วนที่เพิ่มเข้ามา ---
				# 1. รอให้ Player โหลดใหม่สักพัก (รอ DOM เปลี่ยน)
				human_pause_long(1.5, 2.5) 

				# 2. สั่งกด Play อีกรอบ (เผื่อ Player หยุดหลังจากเปลี่ยนช่อง)
				print("    -> Re-clicking Play button...")
				click_media_play_button(driver, selectors, timeout=5)

				# 3. สั่ง Skip Ads ใหม่อีกรอบ (เพราะโฆษณาอาจจะมาใหม่หลังเปลี่ยน source)
				print("    -> Re-skipping Ads...")
				handle_skip_ads_dooball(driver, selectors, rounds=2)
				# -----------------------

				# 4. รอให้ Network วิ่ง (เพิ่มเวลาเล็กน้อยเพื่อให้ request m3u8 ออกไป)
				human_pause_long(3.0, 5.0)

				# collect network
				current = capture_network(driver)
				new = 0
				for u in current:
					if u not in already_found_links:
						already_found_links.add(u)
						new += 1
				print(f"    -> found {new} new m3u8")
			except Exception as ex:
				print(f"    [warn] refresh click failed: {ex}")
			human_pause(0.2, 0.6)
		# end buttons loop
		print(f"  waiting {delay}s before next refresh round...")
		time.sleep(delay)
	print("[refresh] finished")

# -----------------------------
# webdriver factory
# -----------------------------
def make_driver(headless: bool = HEADLESS):
	# install chromedriver binary once (safe to call every worker)
	chromedriver_autoinstaller.install()
	options = webdriver.ChromeOptions()
	options.add_argument("--disable-blink-features=AutomationControlled")
	options.add_argument("--no-sandbox")
	options.add_argument("--disable-dev-shm-usage")
	options.add_experimental_option("excludeSwitches", ["enable-logging"])
	options.add_argument("--log-level=3")
	if headless:
		options.add_argument("--headless=new")
		options.add_argument("--window-size=1366,768")
	driver = webdriver.Chrome(seleniumwire_options={}, options=options)
	# sniff only .m3u8 for speed (selenium-wire scopes)
	if M3U8_ONLY_SCOPES:
		driver.scopes = [r".*\.m3u8(\?.*)?$"]
	else:
		driver.scopes = [".*"]
	return driver

# -----------------------------
# worker: single visit (used by ThreadPoolExecutor)
# -----------------------------
def scan_visit(site: str, selectors: dict, is_dooball: bool) -> Tuple[str, Set[str]]:
	"""
	Performs one visit for a site.
	If is_dooball True, the visit will also run refresh loop (multiple rounds) to collect variations.
	Returns (site, set_of_found_m3u8).
	"""
	found_set: Set[str] = set()
	driver = None
	try:
		print(f"[visit start] {site}")
		driver = make_driver()
		print(f"[visit driver ready] {site}")

		driver.get(site)
		# initial small wait
		human_pause_long(1.0, 2.2)

		# Try to center player (iframe/video) if exists (best-effort)
		try:
			# try safe center by switching to likely iframe then clicking body to activate player
			f = try_switch_to_any_iframe(driver)
			if f:
				try:
					body = driver.find_element(By.TAG_NAME, "body")
					ActionChains(driver).move_to_element(body).click().perform()
				except Exception:
					pass
				switch_back_to_default(driver)
		except Exception:
			pass

		# -----------------------------
		# 🔥 ACTIVATE PLAYER (IMPORTANT)
		# -----------------------------
		activate_player(driver)

		# ▶️ try play
		click_media_play_button(driver, selectors, timeout=10)

		# ⏳ wait for ad DOM to appear
		time.sleep(1.5)

		# ⏭ skip ads (dooball only)
		if is_dooball:
			print("[dooball] aggressive skip ads")
			handle_skip_ads_dooball(driver, selectors, rounds=3)

			print("[dooball] ensure stream start (IMPORTANT)")
			ensure_stream_start(driver)   # ⭐⭐⭐
			human_pause_long(1.2, 2.0)
		else:
			handle_skip_ads(driver, selectors)

		# short wait to allow network m3u8 to appear
		print("[wait] waiting for player to load...")
		for i in range(10):  # รอ 10 รอบ (รวมประมาณ 10–15 วินาที)
			time.sleep(random.uniform(1.0, 1.6))  # รอรายวินาที
			current = capture_network(driver)
			new = 0
			for u in current:
				if u not in found_set:
					found_set.add(u)
					new += 1
			print(f"   [net] +{new} new (round {i+1}/10)")

		# if dooball -> run refresh loop to get variations
		if is_dooball:
			print("[dooball] re-trigger skip before refresh")
			activate_player(driver) # กระตุ้น iframe อีกรอบ
			handle_skip_ads_dooball(driver, selectors, rounds=2) # skip ads อีกรอบ

			# *** CALL THE MODIFIED REFRESH FUNCTION ***
			click_refresh_channels(driver, selectors, found_set, rounds=6, delay=3)

		# final capture
		human_pause(0.8, 1.6)
		final = capture_network(driver)
		for u in final:
			found_set.add(u)

	except Exception as e:
		print(f"[error][visit] {site}: {e}")
	finally:
		try:
			if driver:
				driver.quit()
		except Exception:
			pass

	return site, found_set

# -----------------------------
# helper: chunk tasks
# -----------------------------
def chunked(lst: List[Tuple[str, bool]], n: int) -> List[List[Tuple[str, bool]]]:
	"""Split list into chunks of size n (last chunk may be smaller)."""
	return [lst[i:i+n] for i in range(0, len(lst), n)]

# -----------------------------
# main: build tasks & run in batched parallel rounds
# -----------------------------
def main(
	sites=SITES,
	visits_per_site=VISITS_PER_SITE,
	selectors_path=SELECTORS_FILE,
	max_workers=MAX_WORKERS
):
	chromedriver_autoinstaller.install()
	selectors = load_selectors(selectors_path)
	results_map: Dict[str, Set[str]] = {s: set() for s in sites}

	for site in sites:
		is_db = "dooball" in site.lower()
		
		# --- LOGIC CHANGE HERE ---
		# ถ้าเจอคำว่า dooball ให้ปรับจำนวนรอบเหลือ 1 ทันที
		# ถ้าไม่ใช่ ให้ใช้ค่าตาม config (visits_per_site)
		if is_db:
			current_site_visits = 1
			print(f"[Config] 'dooball' detected for {site} -> Limiting to 1 round.")
		else:
			current_site_visits = visits_per_site
		# -------------------------

		print(f"\n==============================")
		print(f"  SCANNING SITE: {site}")
		print(f"==============================")

		# ใช้ current_site_visits แทน visits_per_site ใน loop นี้
		for round_index in range(1, current_site_visits + 1):
			print(f"[ROUND {round_index}/{current_site_visits}] Launching {max_workers} browsers...")
			tasks = [(site, is_db) for _ in range(max_workers)]

			with ThreadPoolExecutor(max_workers=max_workers) as ex:
				futures = {ex.submit(scan_visit, s, selectors, db): s for s, db in tasks}
				for fut in as_completed(futures):
					site_key = futures[fut]
					try:
						_, found = fut.result()
						results_map[site_key].update(found)
						print(f"[OK] {site_key} → +{len(found)} items")
					except Exception as e:
						print(f"[ERROR] Worker failed: {e}")

			print(f"[ROUND {round_index}] finished for {site}")
			human_pause_long(1.0, 2.0)

	# export excel หลังจบรอบทั้งหมด
	export_xlsx(results_map)

if __name__ == "__main__":
	main()
